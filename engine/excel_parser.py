"""Excel 模板读写 + 校验"""
import io
import re
from dataclasses import dataclass
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 模板列定义 ──
TEMPLATE_COLUMNS = ["序号", "公司名称", "收件人邮箱", "邮件主题"]
TEMPLATE_FILE = "resume_mailer_template.xlsx"

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
header_fill = PatternFill(start_color="1A4B8C", end_color="1A4B8C", fill_type="solid")
header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
body_font = Font(name="微软雅黑", size=10)


def generate_template() -> bytes:
    """生成 Excel 模板文件（内存中）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收件人列表"

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 45

    # 表头
    for col_idx, name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 范例行
    example = [1, "示例公司", "hr@example.com", "实习申请-姓名-学校-专业"]
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = body_font
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left", vertical="center")
        cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@dataclass
class Recipient:
    seq: int
    company: str
    email: str
    subject: str

@dataclass
class ParseResult:
    recipients: List[Recipient]
    errors: List[str]


def parse_workbook(file_bytes: bytes, filename: str = "") -> ParseResult:
    """解析用户上传的 Excel，返回收件人列表和错误信息"""
    errors = []
    recipients = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return ParseResult([], [f"无法打开Excel文件: {e}"])

    ws = wb.active
    if ws is None:
        return ParseResult([], ["Excel文件中没有工作表"])

    rows = list(ws.iter_rows(values_only=False))
    if len(rows) < 2:
        return ParseResult([], ["Excel文件中没有数据行（至少需要表头+1行数据）"])

    # 检测表头
    header_row = rows[0]
    col_map = {}  # column_name -> index (0-based)
    for idx, cell in enumerate(header_row):
        if cell.value is not None:
            val = str(cell.value).strip()
            for template_name in TEMPLATE_COLUMNS:
                if template_name in val or val in template_name:
                    col_map[template_name] = idx

    # 校验必要列
    missing = [c for c in TEMPLATE_COLUMNS if c not in col_map]
    # 允许"序号"缺失（自动生成）
    if "序号" in missing:
        missing.remove("序号")
    if missing:
        return ParseResult([], [f"缺少必要列: {', '.join(missing)}。支持的列名: {', '.join(TEMPLATE_COLUMNS)}"])

    for row_idx, row in enumerate(rows[1:], 2):
        company_cell = row[col_map["公司名称"]]
        email_cell = row[col_map["收件人邮箱"]]
        subject_cell = row[col_map["邮件主题"]]

        company = str(company_cell.value or "").strip()
        email_raw = str(email_cell.value or "").strip()
        subject = str(subject_cell.value or "").strip()

        # 跳过完全空行
        if not company and not email_raw:
            continue

        # 校验邮箱
        email_raw = email_raw.replace("、", ",")
        emails = [e.strip() for e in email_raw.split(",") if e.strip()]
        valid_emails = []
        for e in emails:
            if re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', e):
                valid_emails.append(e)
            else:
                errors.append(f"第{row_idx}行: 邮箱「{e}」格式不正确，已跳过")

        if not valid_emails:
            errors.append(f"第{row_idx}行: 公司「{company}」无有效邮箱，已跳过")
            continue

        # 公司名不能为空
        if not company:
            errors.append(f"第{row_idx}行: 公司名称为空，已跳过")
            continue

        # 序号
        seq_cell = row[col_map.get("序号", -2)]
        if seq_cell and seq_cell.value is not None:
            try:
                seq = int(seq_cell.value)
            except (ValueError, TypeError):
                seq = row_idx - 1
        else:
            seq = len(recipients) + 1

        # 多收件人用逗号分隔
        recipients.append(Recipient(
            seq=seq,
            company=company,
            email=",".join(valid_emails),
            subject=subject,
        ))

    return ParseResult(recipients, errors)


def detect_placeholders(subject: str) -> List[str]:
    """检测主题中的占位符（如 {公司名} {姓名} {岗位}）"""
    return re.findall(r'\{([^}]+)\}', subject)
